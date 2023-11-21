import json
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
from pathlib import Path
from statistics import mean, median
from collections import Counter

def creaexcel():
    libro = Workbook()
    libro.remove(libro.active)  # Elimina la hoja activa predeterminada
    hoja_sheet = libro.create_sheet("Sheet", 0)
    hoja_tipos = libro.create_sheet("Hoja Tipos", 1)
    hoja_colores = libro.create_sheet("Hoja Colores", 2)
    hoja_stats = libro.create_sheet("Hoja de Stats", 3)
    hoja_stats = libro.create_sheet("Hoja Stats Totales", 4)
    #print(libro.sheetnames)
    libro.save(pathxls)

def cargar_datos_pokemon():
    consultajson = Path(__file__).parent.parent / 'Consulta de API' / 'pokemones_[13-11_2.14].json'
    with open(consultajson, "r", encoding="utf-8") as archivo_json:
        datos_pokemones = json.load(archivo_json)
    return datos_pokemones

def infotipos(datos_pokemones):
    typestot = list()
    nomtyptot = {'grass': 0, 'fire': 0, 'water': 0, 'normal': 0, 'fighting': 0, 'flying': 0, 'poison': 0, 'electric': 0,
                 'ground': 0, 'fairy': 0, 'psychic': 0, 'rock': 0, 'steel': 0, 'ice': 0, 'ghost': 0, 'dragon': 0, 'bug': 0}

    for i in datos_pokemones:
        types = i['Tipo/s']
        for j in types:
            typestot.append(j)

    for i in nomtyptot:
        nomtyptot[i] = typestot.count(i)

    try:
        libro = load_workbook(pathxls)
    except FileNotFoundError:
        print("El archivo no existe, lo volveremos a crear, aguarde...")
        creaexcel()

    if len(libro.sheetnames) > 1:
        libro.active = 1
        hoja = libro.active
    else:
        hoja = libro.active

    hoja["A1"] = "Tipo"
    hoja["B1"] = "Cantidad de pokemon"

    count = 2
    for x, y in nomtyptot.items():
        hoja.cell(count, 1, x)
        hoja.cell(count, 2, y)
        count += 1

    libro.save(pathxls)


def infocolor(datos_pokemones):
    colorstot = list()
    nomcolortot = {'red': 0, 'blue': 0, 'yellow': 0, 'green': 0, 'black': 0, 'brown': 0, 'purple': 0, 'gray': 0,
                   'white': 0, 'pink': 0}

    for i in datos_pokemones:
        colorstot.append(i['Tipo de Color'])

    for i in nomcolortot:
        nomcolortot[i] = colorstot.count(i)

    try:
        libro = load_workbook(pathxls)
    except FileNotFoundError:
        print("El archivo no existe, lo volveremos a crear, aguarde...")
        creaexcel()

    if len(libro.sheetnames) > 2:
        libro.active = 2
        hoja = libro.active
    else:
        hoja = libro.active

    hoja["A1"] = "Color"
    hoja["B1"] = "Cantidad de pokemon"

    count = 2
    for x, y in nomcolortot.items():
        hoja.cell(count, 1, x)
        hoja.cell(count, 2, y)
        count += 1

    libro.save(pathxls)

def infostatsid(pokemon):
    if pokemon:
        nomstattot = {'Base Hp': 0, 'Base Attack': 0, 'Base Defense': 0, 'Base Special Attack': 0,
                    'Base Special Defense': 0, 'Base Speed': 0}

        nombre_poke = pokemon['Nombre']
        nombre_poke=nombre_poke.title()
        for x, y in pokemon['Estadísticas'].items():
            for i in nomstattot:
                if i == x:
                    nomstattot[i] = y

        try:
            libro = load_workbook(pathxls)
        except FileNotFoundError:
            print("El archivo no existe, lo volveremos a crear, aguarde...")
            creaexcel()

        if len(libro.sheetnames) > 3:
            libro.active = 3
            hoja = libro.active
        else:
            hoja = libro.active

        hoja.cell(1, 1, 'Nombre')
        hoja.cell(2, 1, nombre_poke)

        count = 2
        for x, y in nomstattot.items():
            hoja.cell(1, count, x)
            hoja.cell(2, count, y)
            count += 1
        # Ajustar el ancho de las columnas según la longitud de los nombres
        for col in hoja.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjuste_width = (max_length + 2)
            hoja.column_dimensions[column].width = adjuste_width

        libro.save(pathxls)
    else:
        print("¡No se encontró el Pokémon!")

def infostatstot(datos_pokemon):
        
    nomstattot = {'Base Hp': list(), 'Base Attack': list(), 'Base Defense': list(), 'Base Special Attack': list(),
                'Base Special Defense': list(), 'Base Speed': list()}
    alltura=list()
    allpeso=list()
    alltypes=list()

    for e in datos_pokemon:
        for x, y in e['Estadísticas'].items():
            if x=='Base Hp':
                nomstattot['Base Hp'].append(y)
            elif x=='Base Attack':
                nomstattot['Base Attack'].append(y)
            elif x=='Base Defense':
                nomstattot['Base Defense'].append(y)
            elif x=='Base Special Attack':
                nomstattot['Base Special Attack'].append(y)
            elif x=='Base Special Defense':
                nomstattot['Base Special Defense'].append(y)
            elif x=='Base Speed':
                nomstattot['Base Speed'].append(y)
    
    for x,y in nomstattot.items():
        nomstattot[x]=int(mean(y))

    for e in datos_pokemon:
        ralt= e['Altura']* 0.1
        alltura.append(ralt)
    medalt=median(alltura)

    for e in datos_pokemon:
        ralt= e['Peso']* 0.1
        allpeso.append(ralt)
    medpeso=median(allpeso)

    for e in datos_pokemon:
        types= e['Tipo/s']
        for j in types:
            alltypes.append(j)
    frectipo=Counter(alltypes)
    mostcomtype=frectipo.most_common()[0]
    smostcomtype=frectipo.most_common()[1]
    tmostcomtype=frectipo.most_common()[2]

    try:
        libro = load_workbook(pathxls)
    except FileNotFoundError:
        print("El archivo no existe, lo volveremos a crear, aguarde...")
        creaexcel()

    if len(libro.sheetnames) > 4:
        libro.active = 4
        hoja = libro.active
    else:
        hoja = libro.active

    hoja["A1"] = "Estadistica"
    hoja["B1"] = "Promedio"

    count = 2
    for x, y in nomstattot.items():
        hoja.cell(count, 1, x)
        hoja.cell(count, 2, y)
        count += 1
    
    hoja["B8"] = "Mediana"
    hoja["A9"] = "Altura (m)"
    hoja["B9"] = medalt
    hoja["A10"] = "Peso (kg)"
    hoja["B10"] = medpeso
    hoja["A11"] = "Top 3 tipos más comunes"
    hoja["A12"] = "Tipos"
    hoja["B12"] = "Frecuencia"
    hoja["A13"] = mostcomtype[0]
    hoja["B13"] = mostcomtype[1]
    hoja["A14"] = smostcomtype[0]
    hoja["B14"] = smostcomtype[1]
    hoja["A15"] = tmostcomtype[0]
    hoja["B15"] = tmostcomtype[1]

    # Ajustar el ancho de las columnas según la longitud de los nombres
    for col in hoja.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjuste_width = (max_length + 2)
        hoja.column_dimensions[column].width = adjuste_width

    libro.save(pathxls)

def mostrarestadistica():
    nstats=list()
    dstats=list()
    conve=dict()
    npes=list()
    dpes=list()
    convepes=dict()
    ntip=list()
    dtip=list()
    convetip=dict()
    try:
        libro = load_workbook(pathxls)
    except FileNotFoundError:
        print("El archivo no existe, adiós!")
        exit()
    
    libro.active = 4
    hoja = libro.active
    
    rango = hoja["A1":"A7"]#Adaptar rango a totalidad de celdas con info
    for celda in rango:
        for objeto in celda:
            if objeto.value==None:
                nstats.append(" ")
            else:
                nstats.append(objeto.value)
    
    rango = hoja["B1":"B7"]#Adaptar rango a totalidad de celdas con info
    for celda in rango:
        for objeto in celda:
            if objeto.value==None:
                dstats.append(" ")
            else:
                dstats.append(objeto.value)
    
    rango = hoja["A8":"A10"]#Adaptar rango a totalidad de celdas con info
    for celda in rango:
        for objeto in celda:
            if objeto.value==None:
                npes.append(" ")
            else:
                npes.append(objeto.value)
    
    rango = hoja["B8":"B10"]#Adaptar rango a totalidad de celdas con info
    for celda in rango:
        for objeto in celda:
            if objeto.value==None:
                dpes.append(" ")
            else:
                dpes.append(objeto.value)

    rango = hoja["A11":"A15"]#Adaptar rango a totalidad de celdas con info
    for celda in rango:
        for objeto in celda:
            if objeto.value==None:
                ntip.append(" ")
            else:
                ntip.append(objeto.value)
    
    rango = hoja["B11":"B15"]#Adaptar rango a totalidad de celdas con info
    for celda in rango:
        for objeto in celda:
            if objeto.value==None:
                dtip.append(" ")
            else:
                dtip.append(objeto.value)        
    
    libro.save(pathxls)
    
    for i in range(len(nstats)):
        conve[nstats[i]]=dstats[i]
    
    for i in range(len(npes)):
        convepes[npes[i]]=dpes[i]
    
    for i in range(len(ntip)):
        convetip[ntip[i]]=dtip[i]
    #print('{0:18}{1:18}'.format("Tipo de Pokémon", "Frecuencia"))
    for tipo, frecuencia in conve.items():
        print('{0:20}{1:5}'.format(tipo, frecuencia))
    print(" ")
    for tipo, frecuencia in convepes.items():
        print('{0:20}{1:5}'.format(tipo, frecuencia))
    print(" ")
    for tipo, frecuencia in convetip.items():
        print('{0:20}{1:5}'.format(tipo, frecuencia))

def graficatipos():
    tipos = list()
    cntipos = list()
    try:
        libro = load_workbook(pathxls)
    except FileNotFoundError:
        print("El archivo no existe, adiós!")
        exit()
    
    libro.active = 1
    hoja = libro.active
    for fila in hoja.values:
        for valor in fila:
            if valor == None:
                continue
            if type(valor) == str:
                tipos.append(valor)
            else:
                cntipos.append(valor)
    tipos.pop(tipos.index('Tipo'))
    tipos.pop(tipos.index('Cantidad de pokemon'))
    ##print(tipos)
    ##print(cntipos)
    libro.save(pathxls)
    x = np.array(tipos)
    y = np.array(cntipos)
    colors = ['#D2DAFF','#B1B2FF','#AAC4FF','#DFCCFB','#D0BFFF','#E5D4FF','#BEADFA','#DCBFFF','#D0A2F7',
            '#6527BE','#9681EB','#E4E4D0','#AEC3AE','#94A684','#EBF3E8','#ADC4CE','#B4BDFF']
    plt.figure(figsize=(13,6))
    plt.bar(x,y, color=colors)
    plt.title("TIPOS DE POKÉMON", fontweight= "bold")
    plt.ylabel('CANTIDAD', fontweight= "bold") 
    plt.xlabel("TIPOS", fontweight= "bold")
    plt.yticks(np.arange(0, 35, 3))
    plt.show()

def graficacolores():
    colores = list()
    cncolores = list()
    try:
        libro = load_workbook(pathxls)
    except FileNotFoundError:
        print("El archivo no existe, adiós!")
        exit()
    
    libro.active = 2
    hoja = libro.active
    for fila in hoja.values:
        for valor in fila:
            if valor == None:
                continue
            if type(valor) == str:
                colores.append(valor)
            else:
                cncolores.append(valor)
    colores.pop(colores.index('Color'))
    colores.pop(colores.index('Cantidad de pokemon'))
    ##print(colores)
    ##print(cncolores)
    libro.save(pathxls)
    y = np.array(cncolores)
    color = ['#EBE3D5','#B0A695','#F1DEC9','#C8B6A6','#A4907C','#8D7B68','#776B5D','#AAA492','#C7BEA2','#9A9483']
    textprops = {"color": "#352F44"}
    plt.figure(figsize=(10,6))
    plt.pie(y,labels=colores, colors=color,autopct='%1.2f%%', textprops=textprops, explode=(0, 0, 0, 0, 0, 0.05, 0, 0, 0, 0))
    plt.title("TIPOS DE COLORES EN POKÉMONES", fontweight= "bold")
    plt.show()
    
def graficastats(pokemon):
    if pokemon:
        nstats = list()
        dstats = list()
        try:
            libro = load_workbook(pathxls)
        except FileNotFoundError:
            print("El archivo no existe, adiós!")
            exit()
        
        libro.active = 3
        hoja = libro.active
        for fila in hoja.values:
            for valor in fila:
                if valor == None:
                    continue
                if type(valor) == str:
                    nstats.append(valor)
                else:
                    dstats.append(valor)
        nstats.pop(nstats.index('Nombre'))
        nombre = nstats.pop(len(nstats) - 1)
        ##print(nstats)
        ##print(dstats)
        libro.save(pathxls)
        x = np.array(nstats)
        y = np.array(dstats)
        linea = "Estadisticas de "+str(nombre)
        colors = ['green','red','blue','grey','orange','deepskyblue','darkmagenta','yellow','brown',
                'hotpink','violet','darkgoldenrod','silver','cyan','indigo','royalblue','black','olivedrab']
        plt.figure(figsize=(13,6))
        plt.barh(x,y, color=colors)
        plt.title(linea, fontweight= "bold")
        plt.xticks(np.arange(0, 105, 5))
        plt.show()
    

pathxls = Path(__file__).parent.parent / 'Reporte' / 'Mi poke.xlsx'
creaexcel()
datos_pokemones = cargar_datos_pokemon()
infotipos(datos_pokemones)
infocolor(datos_pokemones)
infostatstot(datos_pokemones)
#infostatsid(datos_pokemones, 5)

